import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side

# Specify the file path and the output directory
file_path = r"C:\Users\Roy Avrahami\OneDrive - Sensorz\Environment services versions report.xlsx"
output_directory = r"C:\Users\Roy Avrahami\OneDrive - Sensorz\Environment services versions report"
output_filename = 'Updated_with_Borders.xlsx'
output_path = f"{output_directory}\\{output_filename}"

# Load the Excel file
df = pd.read_excel(file_path)

# Process 'Main Test Cases' column
new_rows = []
for index, row in df.iterrows():
    # Split 'Main Test Cases' into separate test cases
    test_cases = str(row['Main Test Cases']).split('Test Name:')
    for i, test_case in enumerate(test_cases):
        if not test_case.strip():
            continue
        new_row = row.copy()
        new_row['Main Test Cases'] = ('Test Name:' + test_case).strip() if i > 0 else test_case.strip()
        new_rows.append(new_row)

# Create a new DataFrame
new_df = pd.DataFrame(new_rows, columns=df.columns)

# Define the border style
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Apply borders using openpyxl
wb = load_workbook(file_path)
ws = wb.active

# Convert the dataframe to rows, and iterate over them
for r_idx, row in enumerate(dataframe_to_rows(new_df, index=False), 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)
        ws.cell(row=r_idx, column=c_idx).border = thin_border

# Attempt to save the workbook with the applied borders
try:
    wb.save(output_path)
    print(f"Modified Excel file with borders saved to: {output_path}")
except PermissionError as e:
    print(f"Failed to save the file due to a permission error: {e}")
except Exception as e:
    print(f"An unexpected error occurred: {e}")
